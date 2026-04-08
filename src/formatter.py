"""
formatter.py
Builds the XLSX (or CSV) report from the enriched, deduplicated articles.
"""
import csv
import io
from datetime import datetime


# Column order and display names for the spreadsheet
_COLUMNS = [
    ("acquirer",        "Acquirer"),
    ("target_company",  "Target Company"),
    ("deal_size",       "Deal Size"),
    ("sector",          "Sector"),
    ("published_date",  "Published Date"),
    ("source",          "Source"),
    ("summary",         "Summary"),
    ("title",           "Headline"),
    ("url",             "URL"),
]


def _sort_articles(articles: list[dict]) -> list[dict]:
    """Sort by acquirer name, then by published date descending."""
    return sorted(
        articles,
        key=lambda a: (
            a.get("acquirer", "").lower(),
            a.get("published_date", ""),
        ),
        reverse=False,
    )


def build_xlsx(articles: list[dict]) -> bytes:
    """
    Build an XLSX file and return it as bytes.
    Uses openpyxl for formatting.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required for XLSX output. Add it to requirements.txt.")

    wb = openpyxl.Workbook()
    ws = wb.active
    run_date = datetime.utcnow().strftime("%Y-%m-%d")
    ws.title = f"Acquisitions {run_date}"

    # ── Header row ────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (_, display_name) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=display_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border

    ws.row_dimensions[1].height = 24

    # ── Data rows ─────────────────────────────────────────────────────────
    alt_fill = PatternFill("solid", fgColor="EEF2F7")

    for row_idx, article in enumerate(_sort_articles(articles), start=2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()

        for col_idx, (field, _) in enumerate(_COLUMNS, start=1):
            value = article.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.border = cell_border
            cell.alignment = Alignment(vertical="top", wrap_text=(field in ("summary", "title")))

            # Make URL a hyperlink
            if field == "url" and value:
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")

    # ── Column widths ─────────────────────────────────────────────────────
    _widths = {
        "acquirer": 18,
        "target_company": 22,
        "deal_size": 14,
        "sector": 18,
        "published_date": 14,
        "source": 20,
        "summary": 45,
        "title": 45,
        "url": 20,
    }
    for col_idx, (field, _) in enumerate(_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _widths.get(field, 16)

    # ── Freeze header row ─────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto-filter ───────────────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_csv(articles: list[dict]) -> str:
    """Build a CSV string from the articles."""
    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=[field for field, _ in _COLUMNS],
        extrasaction="ignore",
    )
    writer.writeheader()
    for article in _sort_articles(articles):
        writer.writerow({field: article.get(field, "") for field, _ in _COLUMNS})
    return output.getvalue()


def format_report(articles: list[dict], output_format: str = "xlsx") -> tuple[bytes | str, str]:
    """
    Build the report and return (data, filename).
    `data` is bytes for xlsx, str for csv.
    """
    run_date = datetime.utcnow().strftime("%Y-%m-%d")

    if output_format.lower() == "csv":
        data = build_csv(articles)
        filename = f"acquisitions_{run_date}.csv"
    else:
        data = build_xlsx(articles)
        filename = f"acquisitions_{run_date}.xlsx"

    return data, filename
